import streamlit as st
import io
import pandas as pd
from openpyxl import load_workbook
from utils import (
    titulo_estilizado,
    deduplicar_colunas,
    copiar_estilo_completo,
    extrair_valor_limpo
)

def renderizar():
    col_v1, col_v2 = st.columns([8, 2])
    with col_v2:
        if st.button("⬅️ Voltar ao Menu", key="btn_voltar_inc"):
            st.session_state["pagina"] = "menu"
            st.rerun()

    titulo_estilizado("INTEGRADOR ==> DADOS GERAIS DO INTERNO >>> SINALE")
    
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("1. Arquivo de ORIGEM")
        source_file = st.file_uploader("Selecione o arquivo de ORIGEM", type=["xlsx", "xls", "csv", "txt"], key="src_upload")
        origem_tem_cabecalho = st.checkbox("Arquivo de Origem tem cabeçalho?", value=True)
    with col2:
        st.subheader("2. Arquivo de DESTINO")
        dest_file = st.file_uploader("Selecione o arquivo de DESTINO (.xlsx)", type=["xlsx"], key="dest_upload")
        header_dest = st.number_input("Linha do cabeçalho no Arquivo de Destino:", value=11, min_value=1)

    if source_file:
        cache_key_src = f"{source_file.name}_{origem_tem_cabecalho}"
        if "source_df" not in st.session_state or st.session_state.get("last_cache_key_src") != cache_key_src:
            hdr = 0 if origem_tem_cabecalho else None
            try:
                source_file.seek(0)
                ext = source_file.name.split('.')[-1].lower()
                engine_util = 'xlrd' if ext == 'xls' else ('openpyxl' if ext == 'xlsx' else None)
                raw = pd.read_excel(source_file, header=hdr, engine=engine_util)
                raw.columns = deduplicar_colunas(raw.columns) if origem_tem_cabecalho else [f"Col {i+1}" for i in range(len(raw.columns))]
                st.session_state["source_df"] = raw
                st.session_state["last_cache_key_src"] = cache_key_src
            except Exception as e:
                st.error(f"Erro ao ler arquivo: {e}")

    if dest_file:
        if "wb_data" not in st.session_state or st.session_state.get("last_dest_name") != dest_file.name:
            dest_file.seek(0)
            st.session_state["wb_data"] = dest_file.getvalue()
            st.session_state["last_dest_name"] = dest_file.name

    df_origem = st.session_state.get("source_df")
    wb_data = st.session_state.get("wb_data")

    if df_origem is not None and wb_data is not None:
        wb = load_workbook(io.BytesIO(wb_data))
        target_sheet = st.selectbox("Escolha a ABA na Planilha de Destino a ser Atualizada:", wb.sheetnames)
        ws = wb[target_sheet]

        st.subheader("3. Seleção de Registros")
        col_busca = st.selectbox("Coluna identificadora (para seleção):", df_origem.columns)
        opcoes_selecao = [f"{val} (Linha {idx})" for idx, val in df_origem[col_busca].items()]
        selected_options = st.multiselect("🔍 Escolha os registros:", opcoes_selecao)
        selected_indices = [int(item.split("(Linha ")[1].replace(")", "")) for item in selected_options]

        if selected_indices:
            st.info(f"📊 **{len(selected_indices)}** registro(s) selecionado(s) para atualização.")

        st.write("---")
        st.subheader("4. Correlação dos dados dos Arquivos ORIGEM X DESTINO")
        mapping = {}
        cols_ui = st.columns(4)
        opcoes_mapeamento = ["--- Não mapear ---", "⚠️ Auto-incrementar (Seq)"] + list(df_origem.columns)
        for i in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_dest, column=i).value
            with cols_ui[(i - 1) % 4]:
                map_val = st.selectbox(f"Col {i} ({header_val or 'S/ Título'})", opcoes_mapeamento, key=f"map_{i}")
                if map_val != "--- Não mapear ---":
                    mapping[i] = map_val

        st.write("---")
        st.subheader("5. Local da Atualização")
        modo_insercao = st.radio("Local de inserção:", ["Final da planilha", "A partir de uma linha específica"])
        target_row = st.number_input("Linha:", min_value=header_dest + 1, value=header_dest + 1) if modo_insercao == "A partir de uma linha específica" else ws.max_row + 1

        st.write("---")
        if st.button("🚀 Processar e Atualizar"):
            if not selected_indices:
                st.error("Selecione itens!")
                st.stop()
            ref_row_idx = (target_row - 1) if modo_insercao == "A partir de uma linha específica" else ws.max_row
            base_seq = 0
            if ref_row_idx >= header_dest:
                val_acima = ws.cell(row=ref_row_idx, column=1).value
                try:
                    base_seq = int(val_acima)
                except:
                    base_seq = 0
            if modo_insercao == "A partir de uma linha específica":
                ws.insert_rows(target_row, amount=len(selected_indices))
            current_row = target_row
            seq_val = base_seq
            for idx in selected_indices:
                seq_val += 1
                ref_row_idx = current_row - 1
                for col_idx in range(1, ws.max_column + 1):
                    target_cell = ws.cell(row=current_row, column=col_idx)
                    ref_cell = ws.cell(row=ref_row_idx, column=col_idx)
                    copiar_estilo_completo(ref_cell, target_cell)
                    if col_idx == 1 or mapping.get(col_idx) == "⚠️ Auto-incrementar (Seq)":
                        target_cell.value = seq_val
                    elif col_idx in mapping:
                        target_cell.value = extrair_valor_limpo(df_origem, idx, mapping[col_idx])
                    else:
                        target_cell.value = ref_cell.value
                current_row += 1
            buffer = io.BytesIO()
            wb.save(buffer)
            st.session_state["wb_data"] = buffer.getvalue()
            st.success("✅ Processamento concluído com sucesso!")
            st.download_button(
                "📥 Baixar Versão Atualizada",
                st.session_state["wb_data"],
                "sinale_atualizado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
